import requests

from bs4 import BeautifulSoup
from openpyxl import Workbook

custom_headers = {
    # 요청 주소 설정하기
    'referer' : 'https://www.daum.net',
    # 웹브라우저 버전명 설정
    'user-agent' : "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36"
}

# 1 ~ 855 회차까지 로또 번호를 수집
# 수집한 내용 엑셀 파일에 저장
# 1 : 15, 2 : 10, 5 : 9, ... <-- 1은 몇 번 나왔고, 2는 몇 번 나왔고를 출력한다.

# Workbook : 파일
# Worksheet : 탭
wb = Workbook()
ws1 = wb.active
ws1.title = 'count'
# ws1.cell(row=1, column=1, value="test") # Excel A1
# = ws1['A1'] = 'aaa'
ws2 = wb.create_sheet(title='list')
# ws2.cell(row=1, column=1, value='test')
# wb.save('test.xlsx')

lotto_numbers = {i:0 for i in range(1, 46)}

for number in range(1, 2):
    url = 'https://search.daum.net/search?w=tot&rtmaxcoll=LOT&DA=LOT&q='+str(number)+'회차%20로또'
    ws1.cell(row=number, column=1, value=f'{number}회차')

    req = requests.get(url, headers=custom_headers)
    html = BeautifulSoup(req.text, "html.parser")
    numbers = html.select('.lottonum .img_lotto')
    for k, n in enumerate(numbers[:6]):
        ws1.cell(row=number, column=k+2, value=int(n.text))
        lotto_numbers[int(n.text)] += +1
    ws1.cell(row=number, column=8, value='보너스 번호:')
    ws1.cell(row=number, column=9, value=f'{numbers[-1].text}')


for k, v in lotto_numbers.items():
    print(k, v)

for row in range(1, 46):
    ws2.cell(row=row, column=1, value=row)
    ws2.cell(row=row, column=2, value=lotto_numbers[row])
wb.save(filename='lotto.xlsx')