import http.server
import socketserver
import io

# pip install xlsxwriter
import xlsxwriter
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

PORT = 8000

class Handler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        #flush 방식 : 이 후 print 문자들은 잡아주고, 이를 변수 또는 파일에 저장하는 것이 가능하다. 이를 이용해 웹 페이지에 print 내용 출력이 가능하다.
        # xlswriter로 만든 데이터를 BytesIO라는 데이터스트림으로 변환해서 출력
        # output = io.BytesIO()
        # workbook = xlsxwriter.Workbook(output, {'in_memory':True}) # 원래는 파일명이 들어와야 하는데 이번엔 변수에 저장
        # worksheet = workbook.add_worksheet()
        # worksheet.write(0,0, "EXCEL TEST")
        # workbook.close()

        # openpyxl 데이터를 BytesIO라는 데이터스트림으로 어떻게 변환해서 출력하는가?
        wb = Workbook()
        ws1 = wb.active
        ws1.cell(row=1, column=1, value="EXCEL TEST")
        output = io.BytesIO(save_virtual_workbook(wb))

        output.seek(0)
        self.send_response(200)
        # Content-Disposition : 컨텐츠를 어떻게 처리할 것인지 정함. # attachment; filename=test.xlsx : 파일을 다운 받는다.
        self.send_header('Content-Disposition','attachment; filename=test.xlsx')
        # Content-type을 엑셀형식(MIME type)으로 제한한다.
        self.send_header('Content-type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # self.send_header가 끝났음을 알린다.
        self.end_headers()
        self.wfile.write(output.read())
        return

print("serving at port", PORT)
httpd = socketserver.TCPServer(('',PORT), Handler)
httpd.serve_forever()